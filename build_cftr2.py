"""Build a REAL, committable CFTR2 extract from the official CFTR2 release xlsx.

Source: data/CFTR2_30January2026.xlsx  (public CFTR2 variant list, cftr2.org).
Output: data/cftr2_2026-01-30.csv

Keeps only variant-level, public columns (no patient-level data).

KEYS — the precise identifier is `cdna_name` (HGVS c., e.g. 'c.1652G>A'), which
covers EVERY variant type (missense, nonsense, splice, indel, synonymous). We also
derive a *1-letter* `protein_variant` key (e.g. 'G551D') from the protein name, but
ONLY as a join convenience for the protein-keyed missense predictors (AlphaMissense/
EVE/ESM1b). It exists for just the simple single-residue missense variants (~780 of
~2,097); every other class (deletions incl. F508del, nonsense like G542X, splice,
etc.) carries an empty protein_variant and must be joined by cdna_name or the
authoritative genomic coordinates (grch38_* below) instead.

REFERENCE BUILD — not assumed: the xlsx header states 'CFTR reference transcript:
NM_000492.4', and the 'Genomic coordinates' sheet ships explicit grch38_* columns
(alongside grch37_*). We read + assert the transcript and record the official header
counts as provenance rather than discarding them. Re-run only when a newer CFTR2
release is dropped into data/.
"""
from pathlib import Path
import re
import openpyxl
import pandas as pd

PKG = Path(__file__).resolve().parent
XLSX = PKG / "data" / "CFTR2_30January2026.xlsx"
OUT = PKG / "data" / "cftr2_2026-01-30.csv"
OUT.parent.mkdir(exist_ok=True)

AA3TO1 = {
    "Ala": "A", "Arg": "R", "Asn": "N", "Asp": "D", "Cys": "C", "Gln": "Q",
    "Glu": "E", "Gly": "G", "His": "H", "Ile": "I", "Leu": "L", "Lys": "K",
    "Met": "M", "Phe": "F", "Pro": "P", "Ser": "S", "Thr": "T", "Trp": "W",
    "Tyr": "Y", "Val": "V",
}
# p.Gly551Asp  -> G551D   (simple single-residue missense only)
MIS = re.compile(r"^p\.([A-Z][a-z]{2})(\d+)([A-Z][a-z]{2})$")


def missense_key(protein_name: str):
    if not protein_name:
        return ""
    m = MIS.match(protein_name.strip())
    if not m:
        return ""
    a, pos, b = m.group(1), m.group(2), m.group(3)
    if a in AA3TO1 and b in AA3TO1:
        return f"{AA3TO1[a]}{pos}{AA3TO1[b]}"
    return ""


wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb["CFTR2 variants by legacy name"]

# --- Provenance from the header rows (1-12), which the data loop below skips. ---
# Capture the release date, official counts, and (critically) the reference
# transcript, so the build documents its own basis instead of assuming GRCh38/MANE.
header_meta = {}
EXPECT_TX = "NM_000492.4"
for r in ws.iter_rows(min_row=1, max_row=12, values_only=True):
    cell = str(r[0]).strip() if r[0] is not None else ""
    if ":" in cell:
        k, v = cell.split(":", 1)
        header_meta[k.strip()] = v.strip()
tx = header_meta.get("CFTR reference transcript", "")
assert EXPECT_TX in tx, (
    f"CFTR2 header transcript is {tx!r}, expected {EXPECT_TX}; the extract's "
    "genomic coordinates + MANE assumptions may no longer hold — check the release.")
print("CFTR2 header provenance:")
for k in ("Date", "Number of patients in CFTR2", "Number of variants reported in CFTR2",
          "Number of variants with interpretations", "CFTR reference transcript"):
    if k in header_meta:
        print(f"  {k}: {header_meta[k]}")
official_reported = header_meta.get("Number of variants reported in CFTR2", "?")

rows = []
for r in ws.iter_rows(min_row=13, values_only=True):
    if r[0] is None:
        continue
    legacy, protein, cdna, alt, alleles, af, prev, cur, changed = r[:9]
    rows.append({
        "protein_variant": missense_key(protein or ""),
        "legacy_name": legacy,
        "protein_name": protein,
        "cdna_name": cdna,
        "cftr2_alleles": alleles,
        "cftr2_af": af,
        "cftr2_class": cur,
    })
df = pd.DataFrame(rows)

# merge GRCh38 genomic coordinates from sheet 2 (on cDNA name)
ws2 = wb["Genomic coordinates"]
g = pd.DataFrame(ws2.iter_rows(min_row=2, values_only=True),
                 columns=[c for c in next(ws2.iter_rows(min_row=1, max_row=1, values_only=True))])
gcols = {"Variant cDNA name": "cdna_name", "grch38_chr": "grch38_chr",
         "grch38_pos": "grch38_pos", "grch38_ref": "grch38_ref", "grch38_alt": "grch38_alt"}
g = g[[c for c in gcols]].rename(columns=gcols).drop_duplicates("cdna_name")
df = df.merge(g, on="cdna_name", how="left")

df.to_csv(OUT, index=False)
print("wrote", OUT.relative_to(PKG), "rows:", len(df))
print("with a 1-letter missense key:", (df["protein_variant"] != "").sum(),
      f"(of {len(df)}; the rest are non-missense — join by cdna_name/genomic coords)")
# Reconcile our row count against CFTR2's own header count (they can differ by a few:
# e.g. rows with alternative-allele notation). Surface it rather than shipping silently.
print(f"row-count check: extracted {len(df)} vs CFTR2 header 'variants reported' = {official_reported}")
print("class counts:\n", df["cftr2_class"].value_counts(dropna=False).to_string())
